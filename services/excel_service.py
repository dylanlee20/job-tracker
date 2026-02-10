from models.job import Job
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from config import Config
import os
import logging

logger = logging.getLogger(__name__)


class ExcelService:
    """Excel 导出服务"""

    @staticmethod
    def export_to_excel(output_path=None):
        """
        导出所有活跃职位到 Excel

        Args:
            output_path: 输出文件路径，默认使用配置文件中的路径

        Returns:
            str: 导出文件的路径
        """
        if output_path is None:
            output_path = Config.EXCEL_EXPORT_PATH

        try:
            # 确保导出目录存在
            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "职位列表"

            # 定义列标题
            headers = [
                '公司', '职位名称', '地点', '发布日期', '截止日期',
                '状态', '首次发现', '最后看到', '最后更新',
                '是否新职位', '是否最近更新', '是否重点', '备注', '来源链接'
            ]

            # 写入标题行
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 获取所有活跃职位
            jobs = Job.query.filter_by(status='active').order_by(Job.first_seen.desc()).all()

            logger.info(f"Exporting {len(jobs)} jobs to Excel...")

            # 写入数据行
            for row_num, job in enumerate(jobs, 2):
                # 数据行
                row_data = [
                    job.company,
                    job.title,
                    job.location,
                    job.post_date.strftime('%Y-%m-%d') if job.post_date else '',
                    job.deadline.strftime('%Y-%m-%d') if job.deadline else '',
                    job.status,
                    job.first_seen.strftime('%Y-%m-%d %H:%M'),
                    job.last_seen.strftime('%Y-%m-%d %H:%M'),
                    job.last_updated.strftime('%Y-%m-%d %H:%M'),
                    '是' if job.is_new else '否',
                    '是' if job.is_updated else '否',
                    '★' if job.is_important else '',
                    job.user_notes or '',
                    job.job_url
                ]

                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value

                    # 设置链接列为超链接
                    if col_num == len(row_data):  # 来源链接列
                        cell.hyperlink = value
                        cell.font = Font(color="0563C1", underline="single")

                # 条件格式：新职位用浅绿色背景
                if job.is_new:
                    for col_num in range(1, len(row_data) + 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

                # 条件格式：更新过的职位用浅橙色背景
                elif job.is_updated:
                    for col_num in range(1, len(row_data) + 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

                # 条件格式：重点职位用黄色星标
                if job.is_important:
                    cell = ws.cell(row=row_num, column=12)  # 是否重点列
                    cell.font = Font(color="FF0000", bold=True, size=14)

            # 调整列宽
            column_widths = {
                'A': 15,  # 公司
                'B': 40,  # 职位名称
                'C': 20,  # 地点
                'D': 12,  # 发布日期
                'E': 12,  # 截止日期
                'F': 10,  # 状态
                'G': 18,  # 首次发现
                'H': 18,  # 最后看到
                'I': 18,  # 最后更新
                'J': 12,  # 是否新职位
                'K': 12,  # 是否最近更新
                'L': 10,  # 是否重点
                'M': 30,  # 备注
                'N': 50   # 来源链接
            }

            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            # 冻结首行
            ws.freeze_panes = 'A2'

            # 添加自动筛选
            ws.auto_filter.ref = f"A1:N{len(jobs) + 1}"

            # 保存文件
            wb.save(output_path)

            logger.info(f"Excel exported successfully to {output_path}")

            return output_path

        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise

    @staticmethod
    def auto_sync_excel():
        """自动同步到默认 Excel 文件"""
        return ExcelService.export_to_excel(Config.EXCEL_EXPORT_PATH)

    @staticmethod
    def export_custom(output_path, filters=None):
        """
        导出带筛选条件的职位到 Excel

        Args:
            output_path: 输出文件路径
            filters: 筛选条件（同 JobService.get_jobs）

        Returns:
            str: 导出文件的路径
        """
        try:
            # 确保导出目录存在
            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "职位列表"

            # 定义列标题
            headers = [
                '公司', '职位名称', '地点', '发布日期', '截止日期',
                '状态', '首次发现', '最后看到', '最后更新',
                '是否新职位', '是否最近更新', '是否重点', '备注', '来源链接'
            ]

            # 写入标题行
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 获取职位（应用筛选条件）
            from services.job_service import JobService
            result = JobService.get_jobs(filters=filters, page=1, per_page=10000)  # 获取所有结果
            jobs_data = result['jobs']

            # 转换为 Job 对象（如果需要）
            jobs = []
            for job_data in jobs_data:
                job = Job.query.get(job_data['id'])
                if job:
                    jobs.append(job)

            logger.info(f"Exporting {len(jobs)} filtered jobs to Excel...")

            # 写入数据行
            for row_num, job in enumerate(jobs, 2):
                row_data = [
                    job.company,
                    job.title,
                    job.location,
                    job.post_date.strftime('%Y-%m-%d') if job.post_date else '',
                    job.deadline.strftime('%Y-%m-%d') if job.deadline else '',
                    job.status,
                    job.first_seen.strftime('%Y-%m-%d %H:%M'),
                    job.last_seen.strftime('%Y-%m-%d %H:%M'),
                    job.last_updated.strftime('%Y-%m-%d %H:%M'),
                    '是' if job.is_new else '否',
                    '是' if job.is_updated else '否',
                    '★' if job.is_important else '',
                    job.user_notes or '',
                    job.job_url
                ]

                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value

                    if col_num == len(row_data):
                        cell.hyperlink = value
                        cell.font = Font(color="0563C1", underline="single")

                if job.is_new:
                    for col_num in range(1, len(row_data) + 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif job.is_updated:
                    for col_num in range(1, len(row_data) + 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

            # 调整列宽和格式（同上）
            column_widths = {
                'A': 15, 'B': 40, 'C': 20, 'D': 12, 'E': 12,
                'F': 10, 'G': 18, 'H': 18, 'I': 18, 'J': 12,
                'K': 12, 'L': 10, 'M': 30, 'N': 50
            }

            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = f"A1:N{len(jobs) + 1}"

            wb.save(output_path)

            logger.info(f"Custom Excel exported successfully to {output_path}")

            return output_path

        except Exception as e:
            logger.error(f"Error exporting custom Excel: {e}")
            raise
